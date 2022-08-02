import openpyxl
import xlsxwriter
from xlsxwriter import Workbook

new_filename = 'C:/Users/dkan9/VScode/EMO_excel/result.xlsx'
#wb_new = Workbook('C:/Users/dkan9/VScode/EMO_excel/Excel/result.xlsx')

from openpyxl import load_workbook
import pandas as pd
from collections import Counter
import os

path = "./Excel"
file_list = os.listdir(path)
print(file_list)


temp_list = [['0'] for _ in range(5)] #분리한 결과 저장하기 위한 빈 리스트 정의
for file_name_raw in file_list:
    file_name = ".\Excel\\" + file_name_raw
    wb = load_workbook(filename=file_name, data_only=True)
    ws = wb.active
    
    category = [r[0].value for r in ws]
    data = [r[1].value for r in ws]

    num = -1
    for d in data: 
        num += 1
        if d.isalpha():
            pass
        else: # 타임스탬프일 때
            temp_list[num].extend(d.split(', ')) # data열에서 data를 콤마기준으로 분리하여 리스트화

category = [r[0].value for r in ws]


print(temp_list)
# 수정#############################################################
for j in range(len(temp_list)):
    for i in range(len(temp_list[j])):
        if ':' in temp_list[j][i]:
            m = temp_list[j][i].split(':')
            s = int(m[0])*60 + int(m[1])
            temp_list[j][i] = s

        else: temp_list[j][i] = int(temp_list[j][i])



# 비슷하면 같은 걸로 만듦
for j in range(len(temp_list)):
    temp_list[j].sort()
    for i in range(len(temp_list[j])-1):
        if temp_list[j][i+1] - temp_list[j][i] <= 5:
            temp_list[j][i+1] = temp_list[j][i]

print(temp_list)


#####################################################################


wb = load_workbook('C:/Users/dkan9/VScode/EMO_excel/result.xlsx')
ws = wb.active
num = 1
# Movie Timestamp Result
for i in range(len(category)-1):
    count_items = Counter(temp_list[i+1]) 
    top_6_items = count_items.most_common(n=6)
    print(top_6_items)
    a_list = []

    for j in top_6_items:
        a_list.append(j[0])
        a_list.sort()
    print(a_list)

    # row = number, column = A, B, C ... 
    ws.cell(column=1, row=num+1, value=category[i+1])
    ws.cell(column=2, row=num+1, value=', '.join(map(str, a_list)))
    # ws.cell(column=2, row=num+1, value=", ".join(a_list))
    num += 1

wb.save(new_filename)

